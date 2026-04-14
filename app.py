#!/usr/bin/env python3
import json, io, os, subprocess, tempfile, zipfile, shutil
from flask import Flask, request, send_file
from flask_cors import CORS
from openpyxl import load_workbook

app = Flask(__name__)
CORS(app)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(BASE_DIR, '2026 ORDER FORM (eng) rev0.xlsx')
if not os.path.exists(TEMPLATE_PATH):
    TEMPLATE_PATH = os.path.join(os.getcwd(), '2026 ORDER FORM (eng) rev0.xlsx')

import logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
logger.info(f"TEMPLATE_PATH: {TEMPLATE_PATH}, exists: {os.path.exists(TEMPLATE_PATH)}")


def fill_template(data):
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    def w(addr, val):
        if val is None or val == '': return
        ws[addr] = val

    # Date — DD/MM/YYYY format (matches original template convention)
    date_val = data.get('date', '')
    if date_val:
        # Convert YYYY/MM/DD or YYYY-MM-DD to DD/MM/YYYY
        parts = date_val.replace('-','/').split('/')
        if len(parts) == 3:
            if len(parts[0]) == 4:  # YYYY/MM/DD
                date_val = f"{parts[2]}/{parts[1]}/{parts[0]}"
            # else already DD/MM/YYYY
        ws['D6'] = date_val
        ws['D6'].data_type = 's'

    w('D7', data.get('fname', ''))
    w('D8', data.get('lname', ''))   # Surname goes in D8, not E8
    w('E9', data.get('yob', ''))
    w('D10', data.get('profession', ''))
    w('D11', data.get('specialty', ''))
    w('D12', data.get('address', ''))
    w('D13', data.get('town', ''))
    w('D14', data.get('country', ''))
    w('D15', data.get('tel', ''))
    w('D16', data.get('email', ''))

    # WD selection
    wd_col = {'300':6,'350':7,'400':8,'450':9,'500':10,'550':11,'600':12,'700':13}
    optic_row = {
        'Galilean|2.0x':20,'Galilean|2.5x':21,'Galilean|3.0x':22,'Galilean|3.5x':23,
        'Prismatic|3.5x':24,'Prismatic|4.0x':25,'Prismatic|5.0x':26,
        'Ergo|3.5x':27,'Ergo|4.5x':28,'Ergo|5.7x':29,
    }
    row = optic_row.get(data.get('optic', ''))
    col = wd_col.get(str(data.get('wd', '')))
    if row and col:
        ws.cell(row=row, column=col).value = '✓'

    # Frame selection: ✓ goes in the cell AFTER the color label
    # Color positions: each color has a label cell, selection box is label_col + 1
    frame_row = {
        'Look':31,'Cool':32,'Techne 2025':33,'Techne [Old]':34,'Techne RX [Old]':35,
        'Techne ASIAN FITTING [OLD]':36,'Techne RX ASIAN FITTING [OLD]':37,
        'Ash 55-17':38,'Ash 53-17':39,'ITA':40,'ITA - Extended Fit':41,'ONE':42,
    }
    # color_name → label_col (✓ goes in label_col + 1)
    frame_color_label_cols = {
        'Look':                          {'Ruby':5,'Emerald':8},
        'Cool':                          {'Ruby':5,'Emerald':8},
        'Techne 2025':                   {'Black/Green':5,'White/Green':8,'Midnight Blue':12,'White/Wisteria':15},
        'Techne [Old]':                  {'White/Red':5,'Black/Green':8,'White/Pink':12,'Black Edition':15},
        'Techne RX [Old]':               {'White/Red':5,'Black/Green':8,'White/Pink':12},
        'Techne ASIAN FITTING [OLD]':    {'White/Red':5,'Black/Green':8,'White/Pink':12,'Black Edition':15},
        'Techne RX ASIAN FITTING [OLD]': {'White/Red':5,'Black/Green':8,'White/Pink':12},
        'Ash 55-17':                     {'Black/Red':5,'Grey/Grey':8,'Purple/Grey':12},
        'Ash 53-17':                     {'Black/Red':5,'Grey/Grey':8,'Purple/Grey':12},
        'ITA':                           {'Midnight Blue':5,'Brown Endura':8,'Green Vespa':12,'Black Edition':15},
        'ITA - Extended Fit':            {'Black Edition':15},
        'ONE':                           {'Desert Sage':5,'Black Edition /Limited Edition':15,'Color Kit':18},
    }
    fr = frame_row.get(data.get('frame', ''))
    if fr:
        color = (data.get('frame_color', '') or '').strip()
        color_map = frame_color_label_cols.get(data.get('frame', ''), {})
        for c_name, label_col in color_map.items():
            if c_name.strip().lower() == color.lower():
                ws.cell(row=fr, column=label_col + 1).value = '✓'  # ✓ in cell after label
                break

    # Customization
    w('E46', data.get('custom_case', ''))
    w('E47', data.get('custom_frame', ''))
    w('N44', data.get('note', ''))

    # Lens type
    lens_row = {'neutral':53,'neutral_cl':55,'distance':57,'intermediate':59,'reading':61,'bifocal':63}
    lr = lens_row.get(data.get('lens_type', ''))
    if lr:
        ws.cell(row=lr, column=9).value = '✓'

    # RX — OD: F(sph) G(cyl) H(axis), OS: J(sph) K(cyl) M(axis)
    rx = data.get('rx', {})
    for dist, rn in [('dist',69),('int',70),('read',71)]:
        for eye, cols in [('od',(6,7,8)),('os',(10,11,13))]:
            for field, cn in zip(('sph','cyl','axis'), cols):
                val = rx.get(f'{eye}_{dist}_{field}')
                if val not in (None, ''):
                    try: val = float(val)
                    except: pass
                    ws.cell(row=rn, column=cn).value = val
    for eye, cn in [('od',6),('os',10)]:
        val = rx.get(f'{eye}_add')
        if val not in (None, ''):
            try: val = float(val)
            except: pass
            ws.cell(row=72, column=cn).value = val

    # Declination
    decl_col = {'18':4,'22':7,'MAX':10}
    dc = decl_col.get(data.get('decl', 'MAX'))
    if dc:
        ws.cell(row=75, column=dc).value = '✓'

    # IPD — Right: F80 G80 H80 / Left: I80 J80 L80 / Total: formula M80 N80 O80
    ipd = data.get('ipd', {})
    for key, addr in [('r1','F80'),('r2','G80'),('r3','H80'),
                      ('l1','I80'),('l2','J80'),('l3','L80')]:
        val = ipd.get(key)
        if val not in (None, ''):
            try: val = float(val)
            except: pass
            ws[addr] = val

    # Headlight
    hl_row = {'LYNX':53,'LYNX PRO':55,'EOS Wireless':57}
    hr = hl_row.get(data.get('headlight', ''))
    if hr:
        ws.cell(row=hr, column=20).value = '✓'

    # Accessories
    acc_map = {
        '701 Overloupes':(69,20),'710 Overloupes':(71,20),
        'Antifog cloth':(71,20),'Case Loupe&Headlight':(73,20),
        'Custom Magnetic Adapter':(63,20),
    }
    for acc in data.get('accessories', []):
        pos = acc_map.get(acc)
        if pos:
            ws.cell(row=pos[0], column=pos[1]).value = '✓'

    # Force recalculation
    wb.calculation.calcMode = 'auto'

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    filled_bytes = buf.getvalue()

    # Re-inject images/drawings — openpyxl drops EMF logo and overwrites drawing rels
    RESTORE_FILES = {
        'xl/media/image1.emf',
        'xl/media/image2.png',
        'xl/drawings/vmlDrawing1.vml',
        'xl/drawings/_rels/drawing1.xml.rels',
    }
    result_buf = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(filled_bytes), 'r') as filled_zip:
        with zipfile.ZipFile(TEMPLATE_PATH, 'r') as orig_zip:
            orig_names = set(orig_zip.namelist())
            filled_names = set(filled_zip.namelist())
            with zipfile.ZipFile(result_buf, 'w', zipfile.ZIP_DEFLATED) as out_zip:
                for item in filled_zip.namelist():
                    if item in RESTORE_FILES and item in orig_names:
                        out_zip.writestr(item, orig_zip.read(item))
                    else:
                        out_zip.writestr(item, filled_zip.read(item))
                for item in orig_names:
                    if item not in filled_names and item in RESTORE_FILES:
                        out_zip.writestr(item, orig_zip.read(item))

    result_buf.seek(0)
    return result_buf.getvalue()


def xlsx_to_pdf(xlsx_bytes):
    tmpdir = tempfile.mkdtemp()
    try:
        xlsx_path = os.path.join(tmpdir, 'order.xlsx')
        with open(xlsx_path, 'wb') as f:
            f.write(xlsx_bytes)
        result = subprocess.run(
            ['libreoffice','--headless','--convert-to','pdf','--outdir',tmpdir,xlsx_path],
            capture_output=True, text=True, timeout=60
        )
        if result.returncode != 0:
            raise RuntimeError(f"LibreOffice: {result.stderr}")
        pdf_path = os.path.join(tmpdir, 'order.pdf')
        if not os.path.exists(pdf_path):
            raise RuntimeError("PDF not generated")
        with open(pdf_path, 'rb') as f:
            return f.read()
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


@app.route('/ping')
def ping():
    return 'ok'


@app.route('/export', methods=['POST','OPTIONS'])
def export():
    if request.method == 'OPTIONS':
        return '', 204
    data = request.get_json()
    fname = data.get('lname') or data.get('fname') or 'Order'
    date = (data.get('date') or '').replace('/','') or 'nodate'
    base = f"Univet_Order_{fname}_{date}"

    xlsx_bytes = fill_template(data)

    try:
        pdf_bytes = xlsx_to_pdf(xlsx_bytes)
    except Exception as e:
        logger.error(f"PDF error: {e}")
        pdf_bytes = None

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"{base}.xlsx", xlsx_bytes)
        if pdf_bytes:
            zf.writestr(f"{base}.pdf", pdf_bytes)
    zip_buf.seek(0)

    return send_file(
        zip_buf,
        mimetype='application/zip',
        as_attachment=True,
        download_name=f"{base}.zip"
    )


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 10000))
    app.run(host='0.0.0.0', port=port)
